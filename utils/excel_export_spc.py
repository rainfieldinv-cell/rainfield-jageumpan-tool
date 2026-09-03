"""
자금판(업무수탁용) 엑셀 만들기 — 예시처럼 '전부 수식' + 예시 서식 재현.

- 시트 3개: 당일자금판 / 이자 스케줄 / 후순위대여
- 숫자를 박아넣지 않고, 입력값만 값으로 두고 나머지는 수식으로 계산.
- 서식: 여백(B2부터 시작), 병합, 표 테두리, 합계행 한 색 통일,
  기초자산(초록)·사모사채(파랑) 색 구별, 두 스케줄을 나란히(같은 행) 배치.
- 이자기간 초일·말일·지급일은 WORKDAY(공휴일 포함)로 주말·공휴일이면 다음 영업일.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from utils.excel_export import (
    FILL_LABEL, FILL_TITLE, FMT_DATE, FMT_RATE, FMT_RATE2, FMT_WON, FONT_B,
    FONT_W, LEFT, _set,
)

# 색 구별용 채우기
FILL_LOAN = PatternFill("solid", fgColor="E2EFDA")   # 기초자산(연초록)
FILL_BOND = PatternFill("solid", fgColor="DDEBF7")   # 사모사채(연파랑)
FILL_SUM = PatternFill("solid", fgColor="D9D9D9")    # 합계(연회색, 한 색)
FILL_HDR = PatternFill("solid", fgColor="F2F2F2")    # 일반 머리글(연회색)
FILL_SEC = PatternFill("solid", fgColor="DCE6F1")    # 개요·유보금 머리글(같은 색)
FILL_CI = PatternFill("solid", fgColor="FCE4D6")     # Cash In 내용(연주황)
FILL_CO = PatternFill("solid", fgColor="E2EFDA")     # Cash Out 내용(연초록)

# 테두리 두께: 표 바깥은 굵게(medium), 안쪽은 얇게(thin) — 예시처럼 구분
_THIN = Side(style="thin", color="B0B0B0")
_MED = Side(style="medium", color="808080")
_THICK = Side(style="thick", color="404040")   # 블록 구분용 굵은 겉 테두리


def _outer_border(ws, rng, side=_THICK):
    """범위의 '겉 테두리만' 굵게 그림(안쪽 셀 선은 그대로 둠). 블록 구분용."""
    rows = ws[rng]
    for c in rows[0]:      # 윗변
        c.border = Border(left=c.border.left, right=c.border.right,
                          top=side, bottom=c.border.bottom)
    for c in rows[-1]:     # 아랫변
        c.border = Border(left=c.border.left, right=c.border.right,
                          top=c.border.top, bottom=side)
    for row in rows:       # 좌·우변
        a, b = row[0], row[-1]
        a.border = Border(left=side, right=a.border.right,
                          top=a.border.top, bottom=a.border.bottom)
        b.border = Border(left=b.border.left, right=side,
                          top=b.border.top, bottom=b.border.bottom)


def _fill_row(ws, row, c1, c2, fill):
    """한 행의 c1~c2 열 전체를 같은 색으로 채움(합계행 색 통일용)."""
    for col in range(c1, c2 + 1):
        ws.cell(row=row, column=col).fill = fill


def _box(ws, rng):
    """범위에 안쪽은 얇은 선, 바깥 테두리는 굵은 선을 넣는다."""
    rows = ws[rng]
    for row in rows:
        for c in row:
            c.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    for c in rows[0]:
        c.border = Border(left=c.border.left, right=c.border.right, top=_MED, bottom=c.border.bottom)
    for c in rows[-1]:
        c.border = Border(left=c.border.left, right=c.border.right, top=c.border.top, bottom=_MED)
    for row in rows:
        a, b = row[0], row[-1]
        a.border = Border(left=_MED, right=a.border.right, top=a.border.top, bottom=a.border.bottom)
        b.border = Border(left=b.border.left, right=_MED, top=b.border.top, bottom=b.border.bottom)


def _align_all(ws):
    """상하 가운데 정렬. 금액(#,##0)은 오른쪽, 나머지는 기존 가로정렬(기본 가운데).
    줄바꿈(wrap)은 끄고, 긴 내용은 열 너비로 한 줄에 보이게 한다."""
    for row in ws.iter_rows():
        for c in row:
            if c.value is None and not c.fill.patternType:
                continue
            nf = c.number_format or ""
            cur = c.alignment
            # 금액(#,##0)만 오른쪽, 나머지는 전부 가운데
            h = "right" if "#,##0" in nf else "center"
            c.alignment = Alignment(horizontal=h, vertical="center",
                                    wrap_text=False)


def build_workbook_spc(plan: dict, accounts: dict) -> BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "당일자금판"
    ws2 = wb.create_sheet("이자 스케줄")
    ws3 = wb.create_sheet("후순위대여")
    for ws in (ws1, ws2, ws3):
        ws.sheet_view.showGridLines = False

    # 사모사채 회차 수(1 또는 2). 2회면 1-2회·추가자산관리수수료를 함께 그림.
    nb = plan.get("nbond", 1)
    has_addfee = bool(plan.get("has_addfee"))
    # 수수료 종류(인수/주선)에 따른 라벨 — 금액·내용 글자가 여기에 맞춰 바뀜
    fee_label = plan.get("fee_label", "인수수수료")        # 인수수수료 / 주선수수료
    fee_rate_label = plan.get("fee_rate_label", "인수수수료율")
    fee_person = plan.get("fee_person", "인수인")          # 인수인 / 주선인

    # 이자 스케줄을 '지급날짜(지급일)' 기준으로 같은 행에 정렬 (발행행 제외)
    # (지급일은 build_schedule 이 이미 주말·공휴일 조정한 값이라 값으로 넣음)
    loan_rows = plan["loan_sch"]["rows"][1:]
    bond_rows = plan["bond_sch"]["rows"][1:]
    bond2_rows = (plan["bond_sch2"]["rows"][1:]
                  if nb == 2 and plan.get("bond_sch2") else [])
    loan_by_pay = {r["지급일"]: r for r in loan_rows}
    bond_by_pay = {r["지급일"]: r for r in bond_rows}
    bond2_by_pay = {r["지급일"]: r for r in bond2_rows}
    pays = sorted(set(loan_by_pay) | set(bond_by_pay) | set(bond2_by_pay))
    SROW = 14                                    # 스케줄 첫 데이터 행
    loan_first_row = SROW + pays.index(loan_rows[0]["지급일"]) if loan_rows else SROW
    bond_first_row = SROW + pays.index(bond_rows[0]["지급일"]) if bond_rows else SROW
    SL = SROW + len(pays) + 2                     # 스케줄 합계 행(빈 여백 2행 포함)
    WHT_TITLE = SL + 2                           # (합계와 1칸 띄우고) 원천세 제목
    WHT_HDR = SL + 3
    WHT_FIRST = SL + 4                           # 원천세 첫 행(=후순위대여)

    # ═══════════ 시트1: 당일자금판 (B2부터, 한 칸씩 여백) ═══════════
    # 긴 내용이 한 줄에 보이도록 열 자체를 넓힘(사채명 G, 수취인 I, 계좌 J, 비고 D 등)
    widths = {"A": 3, "B": 18, "C": 18, "D": 26, "E": 17, "F": 15,
              "G": 32, "H": 15, "I": 26, "J": 26, "K": 16}
    if nb == 2:
        # 1-2회 개요를 J:L에 두는데, J·K는 Cash Out의 수취계좌·비고 열과도 겹침.
        # 수취계좌(J)·사채명 내용(K)이 안 잘리게 넉넉히(1회처럼) 넓힘.
        widths.update({"J": 26, "K": 32, "L": 15})
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w

    ws1.merge_cells("B2:D2")
    _set(ws1, "B2", f"▣ {plan['spc_name']} 사모사채 대출개요 및 자금판",
         font=FONT_B, fill=FILL_LABEL, align=LEFT)

    # 1. 개요 (당일자금판에서는 기초자산·사모사채 개요를 같은 색으로)
    bond1_label = "1-1. 사모사채" if nb == 2 else "1-1. 1회 사모사채"
    ws1.merge_cells("B4:D4"); _set(ws1, "B4", "1. 기초자산(대출) 개요", font=FONT_B, fill=FILL_SEC, align=LEFT)
    ws1.merge_cells("F4:H4"); _set(ws1, "F4", bond1_label, font=FONT_B, fill=FILL_SEC, align=LEFT)
    for c, t in [("B5", "구분"), ("C5", "내용"), ("D5", "비고"),
                 ("F5", "구분"), ("G5", "내용"), ("H5", "비고")]:
        _set(ws1, c, t, font=FONT_B, fill=FILL_SEC)

    # 대출일·만기일은 대출(loan) 날짜, 발행일·만기일은 사모(bond) 날짜 — 다를 수 있어 각각 씀
    left = [("차주명", plan["borrower"], "@"), ("대출금액(원)", plan["loan_amount"], FMT_WON),
            ("대출일", plan["loan_date"], FMT_DATE), ("만기일", plan["loan_maturity"], FMT_DATE),
            ("대출기간(일)", "=C9-C8", FMT_WON), ("금리(연,고정)", plan["loan_rate"], FMT_RATE),
            ("참여수수료", plan["part_rate"], FMT_RATE2)]
    for i, (name, val, fmt) in enumerate(left):
        r = 6 + i
        _set(ws1, f"B{r}", name, font=FONT_B, fill=FILL_LABEL)
        _set(ws1, f"C{r}", val, numfmt=fmt)
    right = [("발행방법", plan["issue_type"], "@"), ("사채명", plan["bond_name"], "@"),
             ("발행금액(원)", plan["issue_amount"], FMT_WON), ("발행일", plan["issue_date"], FMT_DATE),
             ("만기일", plan["bond_maturity"], FMT_DATE), ("대출기간(일)", "=G10-G9", FMT_WON),
             ("금리(연,고정)", plan["issue_rate"], FMT_RATE), (fee_label, plan["uw_fee_rate"], FMT_RATE2)]
    for i, (name, val, fmt) in enumerate(right):
        r = 6 + i
        _set(ws1, f"F{r}", name, font=FONT_B, fill=FILL_LABEL)
        _set(ws1, f"G{r}", val, numfmt=fmt)
    # 기초자산 표에만 '기타' 행 추가 (사모사채엔 없음)
    _set(ws1, "B13", "기타", font=FONT_B, fill=FILL_LABEL)
    ws1.merge_cells("C13:D13")
    _set(ws1, "C13", "후순위 대여는 매 이자지급일에 차주에게 지급 받음", align=LEFT)
    _box(ws1, "B4:D13")
    _box(ws1, "F4:H13")

    # 1-2회 사모사채 개요 (회차 2일 때만, J:L). 발행일·만기는 1-1회(사모 G9/G10)와 같음.
    if nb == 2:
        ws1.merge_cells("J4:L4")
        _set(ws1, "J4", "1-2. 사모사채", font=FONT_B, fill=FILL_SEC, align=LEFT)
        for c, t in [("J5", "구분"), ("K5", "내용"), ("L5", "비고")]:
            _set(ws1, c, t, font=FONT_B, fill=FILL_SEC)
        right2 = [("발행방법", plan["issue_type2"], "@"), ("사채명", plan["bond_name2"], "@"),
                  ("발행금액(원)", plan["issue_amount2"], FMT_WON), ("발행일", "=G9", FMT_DATE),
                  ("만기일", "=G10", FMT_DATE), ("대출기간(일)", "=K10-K9", FMT_WON),
                  ("금리(연,고정)", plan["issue_rate2"], FMT_RATE),
                  (fee_label, plan["uw_fee_rate2"], FMT_RATE2)]
        for i, (name, val, fmt) in enumerate(right2):
            r = 6 + i
            _set(ws1, f"J{r}", name, font=FONT_B, fill=FILL_LABEL)
            _set(ws1, f"K{r}", val, numfmt=fmt)
        _box(ws1, "J4:L13")

    # 2. 자금판
    ws1.merge_cells("B15:D15")
    _set(ws1, "B15", "2. 자금판 (SPC 기준, 단위: 원)", font=FONT_B, align=LEFT)
    _set(ws1, "B16", "집행일", font=FONT_B, fill=FILL_LABEL); _set(ws1, "C16", "=C8", numfmt=FMT_DATE)
    _box(ws1, "B16:C16")  # 집행일 칸 테두리

    ws1.merge_cells("B17:D17"); _set(ws1, "B17", "Cash In", font=FONT_W, fill=FILL_TITLE)
    ws1.merge_cells("E17:K17"); _set(ws1, "E17", "Cash Out", font=FONT_W, fill=FILL_TITLE)
    for c, t in [("B18", "내용"), ("C18", "금액"), ("D18", "지급인"),
                 ("E18", "내용"), ("F18", "금액(공급가)"), ("G18", "부가세"),
                 ("H18", "합계"), ("I18", "수취인"), ("J18", "수취계좌"), ("K18", "비고")]:
        _set(ws1, c, t, font=FONT_B, fill=FILL_HDR)

    # Cash In (B내용 C금액 D지급인) — 인수대금은 회차 합(2회면 G8+K8)
    issue_ref = "=G8+K8" if nb == 2 else "=G8"
    ci = [("사모사채 인수대금", issue_ref, plan["uw_manager"]),
          ("참여수수료", "=C12*C7", plan["borrower"]),
          ("유동화비용", plan["liq"], plan["borrower"]),
          ("선취이자", f"='이자 스케줄'!I{loan_first_row}", plan["borrower"]),
          ("후순위대여", f"='이자 스케줄'!F{WHT_FIRST}", plan["borrower"])]
    for i, (name, val, who) in enumerate(ci):
        r = 19 + i
        # 지급인(D열)은 자동으로 채우지 않고 비워둔다(사용자 요청 — 직접 입력).
        _set(ws1, f"B{r}", name); _set(ws1, f"C{r}", val, numfmt=FMT_WON); _set(ws1, f"D{r}", None)

    # Cash Out (E내용 F공급가 G부가세 H합계 I수취인 J수취계좌)
    def out(r, name, supply, vat, total, who, acc):
        _set(ws1, f"E{r}", name); _set(ws1, f"F{r}", supply, numfmt=FMT_WON)
        _set(ws1, f"G{r}", vat, numfmt=FMT_WON); _set(ws1, f"H{r}", total, numfmt=FMT_WON)
        _set(ws1, f"I{r}", who); _set(ws1, f"J{r}", acc or None, align=LEFT)
    # 19행: 완전히 빈 칸(라벨·금액 모두 없음, 박스·색만 유지 → '1칸 띄움'). 사용자 요청.
    # 순서: (빈 칸) → 자산관리수수료 → 회계법인수수료 → 업무위탁수수료 → 인수수수료
    _set(ws1, "E20", "자산관리수수료"); _set(ws1, "H20", plan["am_total"], numfmt=FMT_WON)
    _set(ws1, "F20", "=ROUND(H20*100/110,0)", numfmt=FMT_WON)
    _set(ws1, "G20", "=ROUND(H20*10/110,0)", numfmt=FMT_WON)
    _set(ws1, "I20", plan["am_manager"]); _set(ws1, "J20", accounts.get("am") or None, align=LEFT)
    out(21, "회계법인수수료", plan["acc_supply"], "=F21*10%", "=F21+G21", plan["acc_manager"], accounts.get("acc"))
    out(22, "업무위탁수수료", plan["bt_supply"], "=F22*10%", "=F22+G22", plan["bt_manager"], accounts.get("bt"))
    # 금액: 직접입력이 있으면 그 값, 없으면 '발행금액×수수료율' 수식
    uw_formula = "=G8*G13+K8*K13" if nb == 2 else "=G8*G13"
    uw_amount = plan["uw_fee_direct"] if plan.get("uw_fee_direct") else uw_formula
    out(23, fee_label, uw_amount, 0, "=F23+G23", plan["uw_manager"], accounts.get("uw"))

    # Cash In / Out 데이터 '행 전체' 색 구별 (예시처럼 그 항목 전체를 칠함)
    for r in range(19, 24):
        for col in range(2, 5):       # B,C,D = Cash In
            ws1.cell(row=r, column=col).fill = FILL_CI
        for col in range(5, 12):      # E~K = Cash Out
            ws1.cell(row=r, column=col).fill = FILL_CO

    # 합계 (26행) — 한 색으로 통일
    _set(ws1, "B26", "합계", font=FONT_B); _set(ws1, "C26", "=SUM(C19:C25)", font=FONT_B, numfmt=FMT_WON)
    _set(ws1, "E26", "합계", font=FONT_B); _set(ws1, "F26", "=SUM(F19:F25)", font=FONT_B, numfmt=FMT_WON)
    _set(ws1, "G26", "=SUM(G19:G25)", font=FONT_B, numfmt=FMT_WON)
    _set(ws1, "H26", "=SUM(H19:H25)", font=FONT_B, numfmt=FMT_WON)
    _fill_row(ws1, 26, 2, 4, FILL_SUM); _fill_row(ws1, 26, 5, 11, FILL_SUM)
    _box(ws1, "B17:D26"); _box(ws1, "E17:K26")

    # 3. 당일 유보금 (다른 표처럼 색 입힘)
    ws1.merge_cells("B28:D28"); _set(ws1, "B28", "3. 당일 유보금", font=FONT_B, fill=FILL_SEC, align=LEFT)
    for c, t in [("B29", "내용"), ("C29", "금액"), ("D29", "비고")]:
        _set(ws1, c, t, font=FONT_B, fill=FILL_SEC)
    _set(ws1, "B30", "후순위대여금"); _set(ws1, "C30", "=C23", numfmt=FMT_WON); _set(ws1, "D30", "매 이자기간 추가 수취")
    _set(ws1, "B31", "예비비"); _set(ws1, "C31", plan["reserve"], numfmt=FMT_WON); _set(ws1, "D31", "등록·등기·이체수수료 등")
    # 지급 유보 이자: 사모 회차 합(2회면 P + W = 1-1회 + 1-2회 첫 이자)
    if nb == 2:
        pay_hold = f"='이자 스케줄'!P{bond_first_row}+'이자 스케줄'!W{bond_first_row}"
    else:
        pay_hold = f"='이자 스케줄'!P{bond_first_row}"
    _set(ws1, "B32", "지급 유보 이자"); _set(ws1, "C32", pay_hold, numfmt=FMT_WON); _set(ws1, "D32", "사모사채 후취 이자")
    if has_addfee:
        # 추가자산관리수수료(기초이자−사모이자) 한 줄 추가 → 검산 정확히 맞음
        _set(ws1, "B33", "추가자산관리수수료")
        _set(ws1, "C33", f"='이자 스케줄'!X{bond_first_row}", numfmt=FMT_WON)
        _set(ws1, "D33", "기초자산 이자 − 사모사채 이자(첫 기간)")
        sum_r = 34
    else:
        sum_r = 33
    _set(ws1, f"B{sum_r}", "합계", font=FONT_B)
    _set(ws1, f"C{sum_r}", f"=SUM(C30:C{sum_r-1})", font=FONT_B, numfmt=FMT_WON)
    # ※ 검산(TRUE/FALSE) 칸은 사용자 요청으로 넣지 않음(대출실행 금액을 비웠기 때문).
    _fill_row(ws1, sum_r, 2, 4, FILL_SUM)
    _box(ws1, f"B28:D{sum_r}")

    # ═══════ 시트2: 이자 스케줄 (지급날짜 기준 정렬, 예시처럼 맨 왼쪽 지급날짜 칸) ═══════
    w2 = {"A": 3, "B": 15, "C": 14, "D": 24, "E": 14, "F": 11, "G": 9,
          "H": 16, "I": 16, "J": 14, "K": 14, "L": 14, "M": 11, "N": 9,
          "O": 16, "P": 16}
    if nb == 2:
        # 1-2회 스케줄(Q~W) + 추가자산관리수수료(X)
        w2.update({"Q": 14, "R": 16, "S": 14, "T": 11, "U": 9,
                   "V": 16, "W": 16, "X": 18})
    for col, w in w2.items():
        ws2.column_dimensions[col].width = w

    def _pay_txt(first, then, kind):
        return (f"{first}개월 {kind}" if first == then
                else f"최초 {first}개월, 이후 {then}개월 {kind}")

    # 입력표 — 기초자산(B:C 라벨/D 값/E:F 비고), 사모사채(J:K 라벨/L 값/M:N 비고)
    _set(ws2, "B2", "기초자산(대출)", font=FONT_B, fill=FILL_LOAN); ws2.merge_cells("B2:D2")
    _set(ws2, "E2", "비고", font=FONT_B, fill=FILL_LOAN); ws2.merge_cells("E2:F2")
    ln = [("대출실행일", "=당일자금판!C8", FMT_DATE), ("차주", "=당일자금판!C6", "@"),
          ("대출금액(원)", "=당일자금판!C7", FMT_WON), ("대출금리", "=당일자금판!C11", FMT_RATE),
          ("참여수수료", "=당일자금판!C12", FMT_RATE2),
          ("이자지급일", _pay_txt(plan["loan_first"], plan["loan_then"], "선취"), "@"),
          ("만기일", "=당일자금판!C9", FMT_DATE)]
    for i, (name, val, fmt) in enumerate(ln):
        r = 3 + i
        ws2.merge_cells(f"B{r}:C{r}"); _set(ws2, f"B{r}", name, font=FONT_B, fill=FILL_LABEL)
        _set(ws2, f"D{r}", val, numfmt=fmt)
        ws2.merge_cells(f"E{r}:F{r}"); _set(ws2, f"E{r}", None)   # 비고 병합
    _box(ws2, "B2:F9")

    bond1_title = "1-1회 사모사채" if nb == 2 else "1회 사모사채"
    _set(ws2, "J2", bond1_title, font=FONT_B, fill=FILL_BOND); ws2.merge_cells("J2:L2")
    _set(ws2, "M2", "비고", font=FONT_B, fill=FILL_BOND); ws2.merge_cells("M2:N2")
    bn = [("사채 발행일", "=당일자금판!G9", FMT_DATE), ("발행유형", "=당일자금판!G6", "@"),
          ("발행금액(원)", "=당일자금판!G8", FMT_WON), ("발행금리", "=당일자금판!G12", FMT_RATE),
          (fee_rate_label, "=당일자금판!G13", FMT_RATE2),
          ("이자지급일", _pay_txt(plan["bond_first"], plan["bond_then"], "후취"), "@"),
          ("만기일", "=당일자금판!G10", FMT_DATE)]
    # 인수인(=인수수수료 수취인)을 '발행금액' 행의 비고에 적음 (예시 파일처럼)
    _uw = (plan.get("uw_manager") or "").strip()
    _uw_memo = f"{fee_person}: {_uw}" if _uw and not _uw.startswith("(") else None
    for i, (name, val, fmt) in enumerate(bn):
        r = 3 + i
        ws2.merge_cells(f"J{r}:K{r}"); _set(ws2, f"J{r}", name, font=FONT_B, fill=FILL_LABEL)
        _set(ws2, f"L{r}", val, numfmt=fmt)
        ws2.merge_cells(f"M{r}:N{r}")
        _set(ws2, f"M{r}", _uw_memo if "발행금액" in name else None, align=LEFT)
    _box(ws2, "J2:N9")

    # 1-2회 사모사채 입력표 (회차 2일 때만, P:T / 값은 R열). 개요(당일자금판 K열) 참조.
    if nb == 2:
        _set(ws2, "P2", "1-2회 사모사채", font=FONT_B, fill=FILL_BOND); ws2.merge_cells("P2:R2")
        _set(ws2, "S2", "비고", font=FONT_B, fill=FILL_BOND); ws2.merge_cells("S2:T2")
        bn2 = [("사채 발행일", "=당일자금판!K9", FMT_DATE), ("발행유형", "=당일자금판!K6", "@"),
               ("발행금액(원)", "=당일자금판!K8", FMT_WON), ("발행금리", "=당일자금판!K12", FMT_RATE),
               (fee_rate_label, "=당일자금판!K13", FMT_RATE2),
               ("이자지급일", _pay_txt(plan["bond2_first"], plan["bond2_then"], "후취"), "@"),
               ("만기일", "=당일자금판!K10", FMT_DATE)]
        for i, (name, val, fmt) in enumerate(bn2):
            r = 3 + i
            ws2.merge_cells(f"P{r}:Q{r}"); _set(ws2, f"P{r}", name, font=FONT_B, fill=FILL_LABEL)
            _set(ws2, f"R{r}", val, numfmt=fmt)
            ws2.merge_cells(f"S{r}:T{r}")
            _set(ws2, f"S{r}", _uw_memo if "발행금액" in name else None, align=LEFT)
        _box(ws2, "P2:T9")

    # 스케줄 머리글 — 맨 왼쪽 '지급날짜'(구분 대용) + 기초(C~I) + 사모(J~P) [+ 1-2회(Q~W) + 추가수수료(X)]
    _set(ws2, "C12", " ▶ 기초자산 이자지급 스케줄 (Cash-in)", font=FONT_B, fill=FILL_LOAN); ws2.merge_cells("C12:I12")
    bond1_sched_title = (" ▶ 1-1회 사모사채 이자지급 스케줄 (Cash-out)" if nb == 2
                         else " ▶ 1회 사모사채 이자지급 스케줄 (Cash-out)")
    _set(ws2, "J12", bond1_sched_title, font=FONT_B, fill=FILL_BOND); ws2.merge_cells("J12:P12")
    _set(ws2, "B13", "지급날짜", font=FONT_B, fill=FILL_HDR)
    lh = ["이자기간(초일)", "이자기간(말일)", "이자지급일(선취)", "일수", "금리(연)", "참여수수료", "이자금액(세전)"]
    for col, t in zip(["C", "D", "E", "F", "G", "H", "I"], lh):
        _set(ws2, f"{col}13", t, font=FONT_B, fill=FILL_LOAN)
    bh = ["이자기간(초일)", "이자기간(말일)", "이자지급일(후취)", "일수", "금리(연)", fee_label, "이자금액(세전)"]
    for col, t in zip(["J", "K", "L", "M", "N", "O", "P"], bh):
        _set(ws2, f"{col}13", t, font=FONT_B, fill=FILL_BOND)
    if nb == 2:
        _set(ws2, "Q12", " ▶ 1-2회 사모사채 이자지급 스케줄 (Cash-out)", font=FONT_B, fill=FILL_BOND)
        ws2.merge_cells("Q12:W12")
        for col, t in zip(["Q", "R", "S", "T", "U", "V", "W"], bh):
            _set(ws2, f"{col}13", t, font=FONT_B, fill=FILL_BOND)
    if has_addfee:
        _set(ws2, "X12", "추가자산관리", font=FONT_B, fill=FILL_SUM)
        _set(ws2, "X13", "추가자산관리수수료", font=FONT_B, fill=FILL_SUM)

    # 데이터: 지급날짜(지급일) 기준 같은 행. 없으면 반대쪽 빈칸.
    loan_first_pay = loan_rows[0]["지급일"] if loan_rows else None
    for idx, pay in enumerate(pays):
        r = SROW + idx
        _set(ws2, f"B{r}", pay, numfmt=FMT_DATE)      # 지급날짜(구분 대용)
        lr = loan_by_pay.get(pay)
        if lr:
            _set(ws2, f"C{r}", lr["초일"], numfmt=FMT_DATE)
            _set(ws2, f"D{r}", lr["말일"], numfmt=FMT_DATE)
            _set(ws2, f"E{r}", lr["지급일"], numfmt=FMT_DATE)
            _set(ws2, f"F{r}", lr["일수"], numfmt=FMT_WON)
            _set(ws2, f"G{r}", "=$D$6", numfmt=FMT_RATE2)
            if pay == loan_first_pay:
                _set(ws2, f"H{r}", "=$D$5*$D$7", numfmt=FMT_WON)
            _set(ws2, f"I{r}", f"=ROUNDDOWN($D$5*G{r}*F{r}/365,0)", numfmt=FMT_WON)
        br = bond_by_pay.get(pay)
        if br:
            _set(ws2, f"J{r}", br["초일"], numfmt=FMT_DATE)
            _set(ws2, f"K{r}", br["말일"], numfmt=FMT_DATE)
            _set(ws2, f"L{r}", br["지급일"], numfmt=FMT_DATE)
            _set(ws2, f"M{r}", br["일수"], numfmt=FMT_WON)
            _set(ws2, f"N{r}", "=$L$6", numfmt=FMT_RATE2)
            _set(ws2, f"P{r}", f"=ROUNDDOWN($L$5*N{r}*M{r}/365,0)", numfmt=FMT_WON)
        if nb == 2:
            br2 = bond2_by_pay.get(pay)
            if br2:
                _set(ws2, f"Q{r}", br2["초일"], numfmt=FMT_DATE)
                _set(ws2, f"R{r}", br2["말일"], numfmt=FMT_DATE)
                _set(ws2, f"S{r}", br2["지급일"], numfmt=FMT_DATE)
                _set(ws2, f"T{r}", br2["일수"], numfmt=FMT_WON)
                _set(ws2, f"U{r}", "=$R$6", numfmt=FMT_RATE2)
                _set(ws2, f"W{r}", f"=ROUNDDOWN($R$5*U{r}*T{r}/365,0)", numfmt=FMT_WON)

    # 인수수수료 금액은 스케줄 맨 위 행에 표시(기초자산 참여수수료 H열과 같은 위치 — 사용자 요청).
    # 사모는 후취라 첫 지급일이 아래 행이지만, 금액 자체는 맨 위(SROW)에 올려 다른 표와 통일.
    _set(ws2, f"O{SROW}", "=$L$5*$L$7", numfmt=FMT_WON)
    if nb == 2:
        _set(ws2, f"V{SROW}", "=$R$5*$R$7", numfmt=FMT_WON)

    # 추가자산관리수수료(X열): 이자기간별 = 기초자산 이자 − 1-1회 이자 − 1-2회 이자.
    # 기초는 선취(지급일=초일 행), 사모는 후취(지급일=말일 행)이라 표시 행이 다르므로,
    # '같은 이자기간 i'끼리 짝지어(기초 i번째 행의 I ↔ 사모 i번째 행의 P·W) 사모 행에 적는다.
    # (지금은 기능 보류 — has_addfee=False 라 그려지지 않음)
    if has_addfee:
        n_fee = min(len(loan_rows), len(bond_rows), len(bond2_rows))
        for i in range(n_fee):
            lrow = SROW + pays.index(loan_rows[i]["지급일"])
            brow = SROW + pays.index(bond_rows[i]["지급일"])
            _set(ws2, f"X{brow}", f"=I{lrow}-P{brow}-W{brow}", numfmt=FMT_WON)

    # 합계 (여백 2행 뒤). 마지막 열: 추가수수료(X=24) > 1-2회(W=23) > 1회만(P=16)
    last_col = 24 if has_addfee else (23 if nb == 2 else 16)
    # '합 계' 라벨: 병합하면 블록 굵은 테두리가 병합셀(C~E)에 안 먹으므로 B칸에만 둠
    _set(ws2, f"B{SL}", "합 계", font=FONT_B)
    _set(ws2, f"F{SL}", f"=SUM(F{SROW}:F{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws2, f"H{SL}", f"=SUM(H{SROW}:H{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws2, f"I{SL}", f"=SUM(I{SROW}:I{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws2, f"M{SL}", f"=SUM(M{SROW}:M{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws2, f"O{SL}", f"=SUM(O{SROW}:O{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws2, f"P{SL}", f"=SUM(P{SROW}:P{SL-1})", font=FONT_B, numfmt=FMT_WON)
    if nb == 2:
        _set(ws2, f"T{SL}", f"=SUM(T{SROW}:T{SL-1})", font=FONT_B, numfmt=FMT_WON)
        _set(ws2, f"V{SL}", f"=SUM(V{SROW}:V{SL-1})", font=FONT_B, numfmt=FMT_WON)
        _set(ws2, f"W{SL}", f"=SUM(W{SROW}:W{SL-1})", font=FONT_B, numfmt=FMT_WON)
    if has_addfee:
        _set(ws2, f"X{SL}", f"=SUM(X{SROW}:X{SL-1})", font=FONT_B, numfmt=FMT_WON)
    _fill_row(ws2, SL, 2, last_col, FILL_SUM)
    _box(ws2, f"B13:{chr(64+last_col)}{SL}")
    # 블록 구분: 기초자산 | 1-1회 사모 | 1-2회 사모 각각 굵은 겉 테두리(제목행 12부터 합계행 SL까지)
    _outer_border(ws2, f"C12:I{SL}")           # 기초자산
    _outer_border(ws2, f"J12:P{SL}")           # 1-1회 사모사채
    if nb == 2:
        _outer_border(ws2, f"Q12:W{SL}")       # 1-2회 사모사채
    if has_addfee:
        _outer_border(ws2, f"X12:X{SL}")       # 추가자산관리수수료

    # 원천세 표 (합계와 1칸 띄우고) — 기초자산 각 이자기간 기준
    _set(ws2, f"B{WHT_TITLE}", "원천세", font=FONT_B)
    for col, t in zip(["B", "C", "D", "E", "F"], ["이자지급일", "이자금액(세전)", "원천세", "지방세", "합계"]):
        _set(ws2, f"{col}{WHT_HDR}", t, font=FONT_B, fill=FILL_HDR)
    loan_period_rows = [SROW + pays.index(lr["지급일"]) for lr in loan_rows]
    for i, pr in enumerate(loan_period_rows):
        r = WHT_HDR + 1 + i
        _set(ws2, f"B{r}", f"=E{pr}", numfmt=FMT_DATE)     # 이자지급일(선취)
        _set(ws2, f"C{r}", f"=I{pr}", numfmt=FMT_WON)      # 이자금액
        _set(ws2, f"D{r}", f"=ROUNDDOWN(C{r}*14%,-1)", numfmt=FMT_WON)
        _set(ws2, f"E{r}", f"=ROUNDDOWN(D{r}*10%,-1)", numfmt=FMT_WON)
        _set(ws2, f"F{r}", f"=D{r}+E{r}", numfmt=FMT_WON)
    wsum = WHT_HDR + 1 + len(loan_rows)
    _set(ws2, f"B{wsum}", "합계", font=FONT_B)
    _set(ws2, f"F{wsum}", f"=SUM(F{WHT_HDR+1}:F{wsum-1})", font=FONT_B, numfmt=FMT_WON)
    _fill_row(ws2, wsum, 2, 6, FILL_SUM)
    _box(ws2, f"B{WHT_HDR}:F{wsum}")

    # ═══════════ 시트3: 후순위대여 ═══════════
    for col, w in {"A": 3, "B": 16, "C": 18, "D": 18, "E": 18}.items():
        ws3.column_dimensions[col].width = w
    _set(ws3, "B2", "■ 후순위 대여", font=FONT_B, align=LEFT)
    for c, t in [("B3", "날짜"), ("C3", "후순위 대여 입금"), ("D3", "후순위 대여 상환"), ("E3", "후순위 대여 잔액")]:
        _set(ws3, c, t, font=FONT_B, fill=FILL_HDR)
    _set(ws3, "B4", "=당일자금판!C16", numfmt=FMT_DATE)
    _set(ws3, "C4", "=당일자금판!C30", numfmt=FMT_WON)
    _set(ws3, "D4", 0, numfmt=FMT_WON)
    _set(ws3, "E4", "=C4-D4", numfmt=FMT_WON)
    _box(ws3, "B3:E4")

    # 정렬: 상하 가운데, 금액은 오른쪽 (세 시트 모두)
    for ws in (ws1, ws2, ws3):
        _align_all(ws)

    # 제목(구획)은 왼쪽 정렬 — 표 안 내용만 가운데, 제목은 원래대로 왼쪽
    _left = Alignment(horizontal="left", vertical="center")
    ws1_lefts = ["B2", "B4", "F4", "B15", "B28"]
    if nb == 2:
        ws1_lefts.append("J4")
    for coord in ws1_lefts:
        ws1[coord].alignment = _left
    ws2_lefts = ["B2", "J2", "C12", "J12", f"B{WHT_TITLE}"]
    if nb == 2:
        ws2_lefts += ["P2", "Q12"]
    for coord in ws2_lefts:
        ws2[coord].alignment = _left
    ws3["B2"].alignment = _left

    wb.calculation.fullCalcOnLoad = True
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
