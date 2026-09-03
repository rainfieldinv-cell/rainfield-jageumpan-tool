"""
최종 엑셀 만들기.

(1) 전체 자금판 — 한 파일 · 탭 2개
        ① 당일자금판    utils/excel_daily2.py     (계약서에서 뽑은 값)
        ② 이자 스케줄   utils/excel_sched2.py     (4단계에서 만든 스케줄 + 후순위대여)
    두 시트는 수식으로 이어져 있다(선취이자·후순위대여·지급 유보 이자).

(2) 사채권자 자금판 — 사모사채 회차별로 한 파일씩
        자금판(후) 의 '이자지급 스케줄' 레이아웃 그대로.
        근거 : 자금판(후)/SPEC_이자스케줄.md  5장
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as _L

from utils.excel_daily2 import write_daily_sheet
from utils.excel_sched2 import write_sched_sheet, write_wht

SHEET_DAILY = "당일자금판"
SHEET_SCHED = "이자 스케줄"


def build_2tab(plan: dict, accounts: dict, sched: dict, info: dict = None) -> BytesIO:
    """전체 자금판 — 탭 2개(당일자금판 · 이자 스케줄)."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = SHEET_DAILY
    ws2 = wb.create_sheet(SHEET_SCHED)

    # ② 이자 스케줄 먼저 그린다 — 당일자금판이 이 시트의 행 번호를 참조하기 때문
    meta = write_sched_sheet(ws2, sched, info)
    wht_meta = None
    if sched.get("wht_on", True) and sched["asset"]:
        wht_meta = write_wht(ws2, meta, sched["asset"],
                             sched.get("wht_rate", 14), sched.get("wht_local", 10))

    # ① 당일자금판
    asset, bonds = sched["asset"], sched["bonds"]
    first_bond = next((b for b in bonds if b["periods"]), None)
    refs = {
        "sheet": SHEET_SCHED,
        "loan_first_row": meta["row_of"][asset[0].pay] if asset else 14,
        "bond_first_row": meta["row_of"][first_bond["periods"][0].pay] if first_bond else 14,
        "wht_first_row": wht_meta["first"] if wht_meta else None,
        "bond_int_cols": meta["bond_int_cols"],
        "addfee_col": meta["addfee_col"],
    }
    write_daily_sheet(ws1, plan, accounts, refs, sched["nbond"])

    wb.calculation.fullCalcOnLoad = True
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ═══════════════════════════════════════════════════════════
# 사채권자 자금판 (회차 하나) — 자금판(후) '이자지급 스케줄' 레이아웃
#   열 : B구분 C초일 D말일 E지급일 F일수 G금리 H인출금액 I이자금액 J인수수수료 K비고(K:M)
#   13행 헤더 / 14행 기준행(-) / 15행~ 이자기간 / 합계 / 각주
# ═══════════════════════════════════════════════════════════
FONT = "맑은 고딕"
_F = Font(name=FONT, size=10)
_FB = Font(name=FONT, size=10, bold=True)
_FW = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
_THIN = Side(style="thin", color="A5A5A5")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")

FILL_GRAY = PatternFill("solid", fgColor="FFF2F2F2")
FILL_TITLE = PatternFill("solid", fgColor="FF808080")
FILL_BLUE = PatternFill("solid", fgColor="FF335693")

MONEY = "#,##0"
DATEF = "yyyy-mm-dd(aaa)"
PCT = '"연" 0.00%'
PCT2 = "0.00%"


def _s(ws, coord, v=None, font=_F, fill=None, align=_CENTER, nf=None, box=True):
    c = ws[coord]
    if v is not None:
        c.value = v
    c.font = font
    c.alignment = align
    if fill:
        c.fill = fill
    if nf:
        c.number_format = nf
    if box:
        c.border = _BOX
    return c


def _fill_rng(ws, c1, r1, c2, r2, fill, white=False):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            if white:
                cell.font = _FW


def build_bond_book(bond: dict, meta: dict) -> BytesIO:
    """
    bond : {"periods": [...], "amount", "rate", "start"}
    meta : {"name","issue_type","issue_date","maturity","fee_mode","fee_rate","fee_amount",
            "pay_type","pay_text","bank","account_no","holder","footnote", 각종 note}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "이자지급 스케줄"
    ws.sheet_view.showGridLines = False

    for col, w in {"B": 14, "C": 16, "D": 16, "E": 19, "F": 13, "G": 9,
                   "H": 15, "I": 15, "J": 15, "K": 12, "L": 12, "M": 18}.items():
        ws.column_dimensions[col].width = w

    ws["D1"] = "=TODAY()"
    ws["D1"].number_format = "mm-dd-yy"

    is_post = meta.get("pay_type", "post") != "pre"

    # ── 상단 정보 ──
    _s(ws, "B2", meta.get("name") or "사모사채", font=_FB)
    _s(ws, "E2", "비고", font=_FB)
    rows = [
        ("사모사채 발행일", meta.get("issue_date"), DATEF, meta.get("issue_date_note")),
        ("발행 유형", meta.get("issue_type"), None, meta.get("issue_type_note")),
        ("사모사채 발행금액(원)", bond["amount"], MONEY, meta.get("amount_note")),
        ("사모사채 발행금리", bond["rate"], PCT, meta.get("rate_note")),
        ("사모사채 인수수수료(원)", None, MONEY, meta.get("fee_note")),
        ("이자지급일", meta.get("pay_text") or ("3개월 후취" if is_post else "3개월 선취"), None, meta.get("pay_note")),
        ("만기일", meta.get("maturity"), DATEF, meta.get("maturity_note")),
    ]
    for i, (label, val, nf, note) in enumerate(rows):
        r = 3 + i
        _s(ws, f"B{r}", label, font=_FB)
        if label.startswith("사모사채 인수수수료"):
            if meta.get("fee_mode") == "amount":
                _s(ws, f"D{r}", meta.get("fee_amount") or 0, nf=MONEY, align=_RIGHT)
            elif meta.get("fee_rate") is not None:
                pct = round(meta["fee_rate"] * 100, 6)
                _s(ws, f"D{r}", f"=D5*{pct}%", nf=MONEY, align=_RIGHT)
            else:
                _s(ws, f"D{r}", nf=MONEY)
        else:
            _s(ws, f"D{r}", val, nf=nf, align=_RIGHT if nf == MONEY else _CENTER)
        _s(ws, f"E{r}", note or None)

    # ── 계좌 ──
    _s(ws, "H2", "사모사채 인수대금 납입 계좌", font=_FB)
    _s(ws, "H5", "은행명", font=_FB)
    _s(ws, "J5", "계좌번호", font=_FB)
    _s(ws, "L5", "예금주", font=_FB)
    _s(ws, "H6", meta.get("bank"))
    _s(ws, "J6", meta.get("account_no"))
    _s(ws, "L6", meta.get("holder"))
    _s(ws, "M11", "(단위 : 원)", align=_RIGHT, box=False)

    # ── 스케줄 ──
    _s(ws, "B12", " ▶ 사모사채 이자지급스케줄", font=_FB, align=_LEFT)
    HEAD = {"B": "구분", "C": "이자기간(초일)", "D": "이자기간(말일)",
            "E": "이자지급일(%s)" % ("후취" if is_post else "선취"),
            "F": "이자계산일수", "G": "금리(연)", "H": "인출금액",
            "I": "이자금액(세전)", "J": "인수수수료", "K": "비고"}
    for col, t in HEAD.items():
        _s(ws, f"{col}13", t, font=_FB)
    for col in ("L", "M"):
        _s(ws, f"{col}13")

    # 기준행 14
    for col in ("B", "C", "D", "F", "G"):
        _s(ws, f"{col}14", "-")
    _s(ws, "E14", "=D3", nf=DATEF)
    _s(ws, "H14", "=D5", nf=MONEY, align=_RIGHT)
    _s(ws, "I14", 0, nf=MONEY, align=_RIGHT)
    _s(ws, "J14", "=D7", nf=MONEY, align=_RIGHT)
    for col in ("K", "L", "M"):
        _s(ws, f"{col}14")

    first = 15
    periods = bond["periods"]
    for i, p in enumerate(periods):
        r = first + i
        _s(ws, f"B{r}", i + 1 if i == 0 else f"=+B{r - 1}+1")
        _s(ws, f"C{r}", p.start, nf=DATEF)
        _s(ws, f"D{r}", p.end, nf=DATEF)
        _s(ws, f"E{r}", p.pay, nf=DATEF)
        _s(ws, f"F{r}", f"=D{r}-C{r}", nf=MONEY)
        _s(ws, f"G{r}", "=$D$6", nf=PCT2)
        _s(ws, f"H{r}", nf=MONEY)
        _s(ws, f"I{r}", f"=ROUNDDOWN($D$5*G{r}*F{r}/365,0)", nf=MONEY, align=_RIGHT)
        _s(ws, f"J{r}", nf=MONEY)
        for col in ("K", "L", "M"):
            _s(ws, f"{col}{r}")
        ws.row_dimensions[r].height = 23.25

    last = first + len(periods) - 1
    tr = last + 1
    _s(ws, f"B{tr}", "합 계", font=_FB)
    for col in ("C", "D", "E"):
        _s(ws, f"{col}{tr}", font=_FB)
    _s(ws, f"F{tr}", f"=SUM(F{first}:F{last})", font=_FB, nf=MONEY)
    _s(ws, f"G{tr}", font=_FB)
    _s(ws, f"H{tr}", f"=SUM(H14:H{last})", font=_FB, nf=MONEY, align=_RIGHT)
    _s(ws, f"I{tr}", f"=SUM(I14:I{last})", font=_FB, nf=MONEY, align=_RIGHT)
    _s(ws, f"J{tr}", f"=SUM(J14:J{last})", font=_FB, nf=MONEY, align=_RIGHT)
    for col in ("K", "L", "M"):
        _s(ws, f"{col}{tr}", font=_FB)
    ws.row_dimensions[tr].height = 23.45

    fr = tr + 1
    _s(ws, f"B{fr}", meta.get("footnote") or "", align=_LEFT, box=False)

    # ── 색 ──
    for c1, r1, c2, r2 in [(2, 2, 4, 2), (5, 2, 6, 2), (8, 2, 13, 3),
                           (8, 5, 9, 5), (10, 5, 11, 5), (12, 5, 13, 5)]:
        _fill_rng(ws, c1, r1, c2, r2, FILL_TITLE, white=True)
    _fill_rng(ws, 2, 3, 3, 9, FILL_GRAY)
    _fill_rng(ws, 2, 13, 13, 13, FILL_GRAY)
    _fill_rng(ws, 2, tr, 13, tr, FILL_GRAY)
    _fill_rng(ws, 2, 12, 13, 12, FILL_BLUE, white=True)

    # ── 병합 ──
    for m in ["B2:D2", "E2:F2", "H2:M3", "B3:C3", "E3:F3", "B4:C4", "E4:F4",
              "B5:C5", "H5:I5", "J5:K5", "L5:M5", "B6:C6", "E6:F6",
              "H6:I7", "J6:K7", "L6:M7", "B7:C7", "E7:F7", "B8:C8", "E8:F8",
              "B9:C9", "E9:F9", "B12:M12"]:
        try:
            ws.merge_cells(m)
        except Exception:
            pass
    try:
        ws.merge_cells(f"B{tr}:E{tr}")
    except Exception:
        pass
    for r in range(13, tr + 1):
        try:
            ws.merge_cells(f"K{r}:M{r}")
        except Exception:
            pass

    wb.calculation.fullCalcOnLoad = True
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
