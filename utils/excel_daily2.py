"""
'당일자금판' 시트 만들기.

utils/excel_export_spc.py 의 시트1(당일자금판)을 그대로 옮기되,
'이자 스케줄' 시트의 행 번호를 밖에서 받는다.
(이자 스케줄 시트를 utils/excel_sched2.py 가 새로 그리므로 행이 달라짐)

구성
  1. 기초자산(대출) 개요 + 사모사채 개요(회차별)
  2. 자금판(SPC 기준) — Cash In / Cash Out
  3. 당일 유보금

'이자 스케줄' 참조
  선취이자       = '이자 스케줄'!I{기초 첫 구간 행}
  후순위대여     = '이자 스케줄'!F{후순위대여 첫 행}
  지급 유보 이자 = '이자 스케줄'!P/W/AD{사모 첫 구간 행}
  추가자산관리수수료 = '이자 스케줄'!{추가수수료 열}{사모 첫 구간 행}
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as _L

FONT = Font(name="맑은 고딕", size=10)
FONT_B = Font(name="맑은 고딕", size=10, bold=True)
FONT_W = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
FONT_BAR = Font(name="맑은 고딕", size=11, bold=True)

FMT_WON = "#,##0"
FMT_RATE = '"연" 0.00%'
FMT_RATE2 = "0.00%"
FMT_DATE = "yyyy-mm-dd(aaa)"

# ── 색 : 회사에서 쓰는 자금판 양식에서 그대로 뽑은 값 ──
#    (다른 테스트용/업무수탁/KT/동교동_아이스리버_자금판_260626.xlsx)
FILL_HEAD = PatternFill("solid", fgColor="2E75B6")   # 개요 헤더 (구분/내용/비고)
FILL_LABEL = PatternFill("solid", fgColor="DEEBF7")  # 개요 라벨 열
FILL_NAVY = PatternFill("solid", fgColor="1F4E79")   # 집행일 · 유보금 헤더
FILL_TITLE = PatternFill("solid", fgColor="D9D9D9")  # Cash In / Cash Out 바
FILL_SEC = None                                      # 섹션 제목은 색 없음
FILL_HDR = PatternFill("solid", fgColor="F2F2F2")    # 열 헤더
FILL_CI = PatternFill("solid", fgColor="FFF2CC")     # Cash In  (연노랑)
FILL_CO = PatternFill("solid", fgColor="E2F0D9")     # Cash Out (연초록)
FILL_SUM = PatternFill("solid", fgColor="F2F2F2")    # 합계
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

_THIN = Side(style="thin", color="B0B0B0")
_MED = Side(style="medium", color="808080")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 사모사채 개요 블록의 시작 열 : F(6) / J(10) / N(14)
BOND_COL = [6, 10, 14]


def _set(ws, coord, value=None, font=FONT, fill=None, align=CENTER, numfmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = font
    c.alignment = align
    if fill:
        c.fill = fill
    if numfmt:
        c.number_format = numfmt
    return c


def _box(ws, c1, r1, c2, r2):
    """안쪽은 얇게, 바깥은 굵게."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(
                top=_MED if r == r1 else _THIN, bottom=_MED if r == r2 else _THIN,
                left=_MED if c == c1 else _THIN, right=_MED if c == c2 else _THIN)


def _fill_row(ws, row, c1, c2, fill):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill


def write_daily_sheet(ws, plan: dict, accounts: dict, refs: dict, nb: int):
    """
    plan     : compute_spc() 결과 (계약서에서 뽑은 값들)
    accounts : {"uw":..., "am":..., "acc":...}  수취계좌
    refs     : {"sheet": "이자 스케줄", "loan_first_row": int, "bond_first_row": int,
                "wht_first_row": int|None, "bond_int_cols": ["P","W","AD"],
                "addfee_col": "X"|None}
    nb       : 사모사채 회차 수 (1~3)
    """
    accounts = accounts or {}
    SH = "'%s'!" % refs["sheet"]
    fee_label = plan.get("fee_label", "인수수수료")
    ws.sheet_view.showGridLines = False

    # 4단계(당일자금판 확인)에서 표에 써 넣은 비고·고친 금액
    D = plan.get("_daily") or {}
    NOTE = D.get("notes") or {}

    def note(group, i, k=None):
        v = NOTE.get(group)
        if k is not None:
            v = (v or [])[k] if v and k < len(v) else None
        if not v or i >= len(v):
            return None
        return (str(v[i]).strip() or None)

    CI_OV = D.get("ci") or []
    CO_OV = D.get("co") or []
    RES_OV = D.get("res") or []

    def ov(lst, i, field=None):
        """4단계에서 직접 써 넣은 값. 없으면 None(=원래 수식 그대로)."""
        if i >= len(lst):
            return None
        v = lst[i]
        if field and isinstance(v, dict):
            return v.get(field)
        return v if not isinstance(v, dict) else None

    # ── 열 폭 ──
    widths = {"A": 3, "B": 18, "C": 18, "D": 26, "E": 17, "F": 15,
              "G": 32, "H": 15, "I": 26, "J": 26, "K": 16}
    if nb >= 2:
        widths.update({"J": 26, "K": 32, "L": 15})
    if nb >= 3:
        widths.update({"M": 3, "N": 15, "O": 32, "P": 15})
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── 제목 ──
    ws.merge_cells("B2:D2")
    _set(ws, "B2", "▣ %s 사모사채 대출개요 및 자금판" % (plan.get("spc_name") or "SPC"),
         font=FONT_B, fill=FILL_LABEL, align=LEFT)

    # ── 1. 기초자산 개요 ──
    ws.merge_cells("B4:D4")
    _set(ws, "B4", "1. 기초자산(대출) 개요", font=FONT_B, align=LEFT)
    for c, t in [("B5", "구분"), ("C5", "내용"), ("D5", "비고")]:
        _set(ws, c, t, font=FONT_W, fill=FILL_HEAD)

    left = [("차주명", plan.get("borrower"), "@"),
            ("대출금액(원)", plan.get("loan_amount"), FMT_WON),
            ("대출일", plan.get("loan_date"), FMT_DATE),
            ("만기일", plan.get("loan_maturity"), FMT_DATE),
            ("대출기간(일)", "=C9-C8", FMT_WON),
            ("금리(연,고정)", plan.get("loan_rate"), FMT_RATE),
            ("참여수수료", plan.get("part_rate"), FMT_RATE2)]
    for i, (name, val, fmt) in enumerate(left):
        r = 6 + i
        _set(ws, f"B{r}", name, font=FONT_B, fill=FILL_LABEL)
        _set(ws, f"C{r}", val, numfmt=fmt)
        _set(ws, f"D{r}", note("asset", i), align=LEFT)
    _set(ws, "B13", "기타", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("C13:D13")
    _set(ws, "C13", D.get("asset_etc") or "후순위 대여는 매 이자지급일에 차주에게 지급 받음",
         align=LEFT)
    _box(ws, 2, 4, 4, 13)

    # ── 1-N. 사모사채 개요 ──
    for k in range(nb):
        c0 = BOND_COL[k]
        L, V, N = _L(c0), _L(c0 + 1), _L(c0 + 2)
        ws.merge_cells(f"{L}4:{N}4")
        _set(ws, f"{L}4",
             ("1-1. 1회 사모사채" if nb == 1 else "1-%d. 사모사채" % (k + 1)),
             font=FONT_B, align=LEFT)
        for c, t in [(f"{L}5", "구분"), (f"{V}5", "내용"), (f"{N}5", "비고")]:
            _set(ws, c, t, font=FONT_W, fill=FILL_HEAD)

        if k == 0:
            itype, bname = plan.get("issue_type"), plan.get("bond_name")
            amt, rate, feerate = plan.get("issue_amount"), plan.get("issue_rate"), plan.get("uw_fee_rate")
        elif k == 1:
            itype, bname = plan.get("issue_type2"), plan.get("bond_name2")
            amt, rate, feerate = plan.get("issue_amount2"), plan.get("issue_rate2"), plan.get("uw_fee_rate2")
        else:
            itype, bname = plan.get("issue_type3"), plan.get("bond_name3")
            amt, rate, feerate = plan.get("issue_amount3"), plan.get("issue_rate3"), plan.get("uw_fee_rate3")

        # 발행일·만기는 1회차와 같으면 그 칸을 참조 (예시와 동일)
        idate = plan.get("issue_date") if k == 0 else "=%s9" % _L(BOND_COL[0] + 1)
        imat = plan.get("bond_maturity") if k == 0 else "=%s10" % _L(BOND_COL[0] + 1)

        rows = [("발행방법", itype, "@"), ("사채명", bname, "@"),
                ("발행금액(원)", amt, FMT_WON), ("발행일", idate, FMT_DATE),
                ("만기일", imat, FMT_DATE),
                ("대출기간(일)", f"={V}10-{V}9", FMT_WON),
                ("금리(연,고정)", rate, FMT_RATE), (fee_label, feerate, FMT_RATE2)]
        for i, (name, val, fmt) in enumerate(rows):
            r = 6 + i
            _set(ws, f"{L}{r}", name, font=FONT_B, fill=FILL_LABEL)
            _set(ws, f"{V}{r}", val, numfmt=fmt)
            _set(ws, f"{N}{r}", note("bond", i, k), align=LEFT)
        _box(ws, c0, 4, c0 + 2, 13)

    # ── 2. 자금판 (SPC 기준) ──
    ws.merge_cells("B15:D15")
    _set(ws, "B15", "2. 자금판 (SPC 기준, 단위: 원)", font=FONT_B, align=LEFT)
    _set(ws, "B16", "집행일", font=FONT_W, fill=FILL_NAVY)
    _set(ws, "C16", "=C8", numfmt=FMT_DATE)
    _box(ws, 2, 16, 3, 16)

    ws.merge_cells("B17:D17")
    _set(ws, "B17", "Cash In", font=FONT_BAR, fill=FILL_TITLE)
    ws.merge_cells("E17:K17")
    _set(ws, "E17", "Cash Out", font=FONT_BAR, fill=FILL_TITLE)
    for c, t in [("B18", "내용"), ("C18", "금액"), ("D18", "지급인"),
                 ("E18", "내용"), ("F18", "금액(공급가)"), ("G18", "부가세"),
                 ("H18", "합계"), ("I18", "수취인"), ("J18", "수취계좌"), ("K18", "비고")]:
        _set(ws, c, t, font=FONT_B, fill=FILL_HDR)

    # Cash In — 4단계 표에서 금액을 직접 쳤으면 그 숫자를, 아니면 원래 수식을 쓴다
    bond_val_cols = [_L(BOND_COL[k] + 1) for k in range(nb)]
    issue_ref = "=" + "+".join("%s8" % c for c in bond_val_cols)
    wht_ref = ("=%sF%d" % (SH, refs["wht_first_row"])) if refs.get("wht_first_row") else None
    ci = [("사모사채 인수대금", issue_ref),
          ("참여수수료", "=C12*C7"),
          ("유동화비용", plan.get("liq")),
          ("선취이자", "=%sI%d" % (SH, refs["loan_first_row"])),
          ("후순위대여", wht_ref)]
    for i, (name, val) in enumerate(ci):
        r = 19 + i
        typed = ov(CI_OV, i, "amount")
        _set(ws, f"B{r}", name)
        _set(ws, f"C{r}", typed if typed is not None else val, numfmt=FMT_WON)
        _set(ws, f"D{r}", ov(CI_OV, i, "payer") or None, align=LEFT)

    # Cash Out — 19행은 대출실행 자리
    def out(r, i, name, supply, vat, total, who, acc):
        s, v, t = ov(CO_OV, i, "supply"), ov(CO_OV, i, "vat"), ov(CO_OV, i, "total")
        _set(ws, f"E{r}", name)
        _set(ws, f"F{r}", s if s is not None else supply, numfmt=FMT_WON)
        _set(ws, f"G{r}", v if v is not None else vat, numfmt=FMT_WON)
        _set(ws, f"H{r}", t if t is not None else total, numfmt=FMT_WON)
        _set(ws, f"I{r}", who or None)
        _set(ws, f"J{r}", acc or None, align=LEFT)
        _set(ws, f"K{r}", note("co", i), align=LEFT)

    out(19, 0, "대출실행", None, None, "=F19+G19",
        plan.get("loan_recipient"), accounts.get("loan"))

    am_total = plan.get("am_total")
    out(20, 1, "자산관리수수료",
        "=ROUND(H20*100/110,0)" if am_total else None,
        "=ROUND(H20*10/110,0)" if am_total else None,
        am_total, plan.get("am_manager"), accounts.get("am"))

    out(21, 2, "회계법인수수료", plan.get("acc_supply"), "=F21*10%", "=F21+G21",
        plan.get("acc_manager"), accounts.get("acc"))
    out(22, 3, "업무위탁수수료", plan.get("bt_supply"), "=F22*10%", "=F22+G22",
        plan.get("bt_manager"), accounts.get("bt"))

    uw_formula = "=" + "+".join("%s8*%s13" % (c, c) for c in bond_val_cols)
    uw_amount = plan.get("uw_fee_direct") or uw_formula
    out(23, 4, fee_label, uw_amount, 0, "=F23+G23",
        plan.get("uw_manager"), accounts.get("uw"))

    for r in range(19, 24):
        _fill_row(ws, r, 2, 4, FILL_CI)
        _fill_row(ws, r, 5, 11, FILL_CO)
    for r in (24, 25):
        _fill_row(ws, r, 2, 11, FILL_WHITE)

    _set(ws, "B26", "합계", font=FONT_B)
    _set(ws, "C26", "=SUM(C19:C25)", font=FONT_B, numfmt=FMT_WON)
    _set(ws, "E26", "합계", font=FONT_B)
    for c in ("F", "G", "H"):
        _set(ws, f"{c}26", f"=SUM({c}19:{c}25)", font=FONT_B, numfmt=FMT_WON)
    _fill_row(ws, 26, 2, 4, FILL_SUM)
    _fill_row(ws, 26, 5, 11, FILL_SUM)
    _box(ws, 2, 17, 4, 26)
    _box(ws, 5, 17, 11, 26)

    # ── 3. 당일 유보금 ──
    ws.merge_cells("B28:D28")
    _set(ws, "B28", "3. 당일 유보금", font=FONT_B, align=LEFT)
    for c, t in [("B29", "내용"), ("C29", "금액"), ("D29", "비고")]:
        _set(ws, c, t, font=FONT_W, fill=FILL_NAVY)

    hold = "=" + "+".join("%s%s%d" % (SH, c, refs["bond_first_row"])
                          for c in refs["bond_int_cols"][:nb])
    res = [("후순위대여금", "=C23", "매 이자기간 추가 수취"),
           ("예비비", plan.get("reserve"), "등록·등기·이체수수료 등"),
           ("지급 유보 이자", hold, "사모사채 후취 이자")]
    if refs.get("addfee_col"):
        res.append(("추가자산관리수수료",
                    "=%s%s%d" % (SH, refs["addfee_col"], refs["bond_first_row"]),
                    "기초자산 이자 − 사모사채 이자(첫 기간)"))
    for i, (name, val, memo) in enumerate(res):
        r = 30 + i
        typed = ov(RES_OV, i)
        _set(ws, f"B{r}", name)
        _set(ws, f"C{r}", typed if typed is not None else val, numfmt=FMT_WON)
        _set(ws, f"D{r}", note("res", i) or memo, align=LEFT)
        _fill_row(ws, r, 2, 4, FILL_WHITE)

    sum_r = 30 + len(res)
    _set(ws, f"B{sum_r}", "합계", font=FONT_B)
    _set(ws, f"C{sum_r}", f"=SUM(C30:C{sum_r - 1})", font=FONT_B, numfmt=FMT_WON)
    _set(ws, f"D{sum_r}", None)
    _fill_row(ws, sum_r, 2, 4, FILL_SUM)
    _box(ws, 2, 28, 4, sum_r)
