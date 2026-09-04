"""
'이자 스케줄' 시트 만들기 — 자금판(후) 서식 그대로.

근거 : 자금판(후)/SPEC_이자스케줄.md  6장 "엑셀 ② 전체(업무수탁) 자금판"

열 배치 (세그먼트당 7열)
    A  여백
    B  지급날짜
    C~I    기초자산          (정보블록 값 열 D · 이자금액 I)
    J~P    사모사채 1-1회    (값 열 L · 이자금액 P)
    Q~W    사모사채 1-2회    (값 열 S · 이자금액 W)
    X~AD   사모사채 1-3회    (값 열 Z · 이자금액 AD)
    그 다음 한 칸  추가자산관리수수료 (1회 Q · 2회 X · 3회 AE)

세그먼트 내부 : 초일 · 말일 · 지급일 · 일수 · 금리 · 수수료 · 이자금액

수식
    일수     = 말일-초일
    금리     = $D$6 / $L$6 / $S$6 / $Z$6
    이자금액 = ROUNDDOWN($D$5*금리셀*일수셀/365,0)
    참여수수료(첫 구간) = $D$5*$D$7
    인수수수료(발행일 행) = $L$7*$L$5 (요율) 또는 $L$7 (직접금액)
"""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as _L

FONT = "맑은 고딕"
MONEY = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
DATEF = "[$-412]yyyy\\/mm\\/dd\\/ddd"
PCT2 = "0.00%"
RATE_INFO = '"연"\\ 0.00%'
NUM = "#,##0"

SEG_W = 7
ASSET_BASE = 3                                   # C
BOND_BASE = [ASSET_BASE + SEG_W * i for i in (1, 2, 3)]   # 10(J) 17(Q) 24(X)

# 색 (자금판(후) 와 같은 값)
F_GRAY = "F2F2F2"
F_ASSET = "2F6B45"
F_BONDS = ["1F4F8A", "1A4677", "13355A"]
F_FEE = "CCC0DA"
C_NAVY = "FF1A2B5E"

_THIN = Side(style="thin")
_MED = Side(style="medium")
_DASH = Side(style="dashed")
_DOT = Side(style="dotted")

ASSET_HEAD = ["이자기간(초일)", "이자기간(말일)", "이자지급일(선취)", "이자계산일수",
              "금리(연)", "참여수수료", "이자금액(세전)"]
BOND_HEAD = ["이자기간(초일)", "이자기간(말일)", "이자지급일(후취)", "이자계산일수",
             "금리(연)", "인수수수료", "이자금액(세전)"]

START_ROW = 14          # 데이터 첫 행
HEAD_G, HEAD_C = 12, 13  # 그룹헤더 / 열헤더


def last_bond_col(nb):
    return BOND_BASE[nb - 1] + SEG_W - 1


def addfee_col(nb):
    return last_bond_col(nb) + 1


def bond_label(idx, nb):
    return "사모사채" if nb == 1 else "1-%d회 사모사채" % idx


def seg_cols(base):
    return {"start": base, "end": base + 1, "pay": base + 2, "days": base + 3,
            "rate": base + 4, "fee": base + 5, "int": base + 6}


def _set(ws, coord, value=None, bold=False, nf=None, h="center", wrap=False,
         fill=None, white=False):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=10, bold=bold,
                  color="FFFFFFFF" if white else None)
    c.alignment = Alignment(horizontal=(None if h == "none" else h),
                            vertical="center", wrap_text=wrap)
    if nf:
        c.number_format = nf
    if fill:
        c.fill = PatternFill("solid", fgColor="FF" + fill)
    return c


def _box(ws, coord):
    ws[coord].border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill_rng(ws, c1, r1, c2, r2, hexv, text=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor="FF" + hexv)
            if text:
                f = cell.font
                cell.font = Font(name=FONT, size=10, bold=True, color=text)


# ─────────────────────────────────────────────
# 상단 정보 블록 (2~9행)
#   라벨 = base(2칸 병합) · 값 = base+2 · 비고 = base+3(2칸 병합)
# ─────────────────────────────────────────────
def _info_block(ws, label_col, val_col, note_col, title, rows):
    """제목·비고 머리는 굵은 상자, 안쪽 줄은 모두 같은 점선으로.

    (예전에는 라벨·값은 dashed, 비고는 dotted 라 줄이 서로 어긋나 보였고
     비고 칸은 오른쪽 선이 없어 상자가 열려 있었다.)
    """
    L, V, N = _L(label_col), _L(val_col), _L(note_col)
    NE = _L(note_col + 1)
    ws.merge_cells(f"{L}2:{V}2")
    t = _set(ws, f"{L}2", title, bold=True)
    t.border = Border(top=_MED, bottom=_MED, left=_MED, right=_MED)
    ws.merge_cells(f"{N}2:{NE}2")
    b = _set(ws, f"{N}2", "비고", bold=True)
    b.border = Border(top=_MED, bottom=_MED, left=_MED, right=_MED)

    last = len(rows) - 1
    for i, r in enumerate(rows):
        row = 3 + i
        close = _MED if i == last else _DASH
        ws.merge_cells(f"{L}{row}:{_L(label_col + 1)}{row}")
        lc = _set(ws, f"{L}{row}", r["label"], bold=True)
        lc.border = Border(top=_DASH, bottom=close, left=_MED, right=_DASH)

        vc = ws[f"{V}{row}"]
        vc.value = r.get("value")
        is_money = r.get("nf") == MONEY
        _set(ws, f"{V}{row}", nf=r.get("nf"), h="none" if is_money else "center")
        vc.border = Border(top=_DASH, bottom=close, left=_DASH, right=_MED)

        ws.merge_cells(f"{N}{row}:{NE}{row}")
        nc = _set(ws, f"{N}{row}", r.get("note") or "", h="center")
        nc.border = Border(top=_DASH, bottom=close, left=_MED, right=_MED)


# ─────────────────────────────────────────────
# 메인
#   sched : utils/step2_ui.render 가 돌려준 결과
#   info  : {"asset_title", "asset_note"..., "bonds":[{title, issue_type, fee_mode, fee_rate, fee_amount, notes}]}
# ─────────────────────────────────────────────
def write_sched_sheet(ws, sched: dict, info: dict = None):
    info = info or {}
    nb = sched["nbond"]
    asset = sched["asset"]
    bonds = sched["bonds"]
    am = sched["asset_meta"]
    use_fee = sched.get("use_addfee") and sched.get("addfee")

    ws.sheet_view.showGridLines = False

    # ── 열 폭 ──
    ASSET_W = [15.5, 29.6, 16.4, 12.9, 8.5, 15.1, 15]
    BOND_W = [15.5, 15.5, 17.8, 12.8, 9.8, 14.2, 15]
    ws.column_dimensions["A"].width = 4.5
    ws.column_dimensions["B"].width = 15
    for i, w in enumerate(ASSET_W):
        ws.column_dimensions[_L(ASSET_BASE + i)].width = w
    for k in range(nb):
        for i, w in enumerate(BOND_W):
            ws.column_dimensions[_L(BOND_BASE[k] + i)].width = w
    if use_fee:
        ws.column_dimensions[_L(addfee_col(nb))].width = 16

    # ── B1 = TODAY() ──
    ws["B1"] = "=TODAY()"
    ws["B1"].number_format = "mm-dd-yy"
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    # ── 정보 블록 ──
    def pay_text(pay_type, rules):
        m = rules[0].months if rules else 3
        return "%d개월 %s" % (m, "선취" if pay_type == "pre" else "후취")

    def _note(lst, i):
        return lst[i] if (lst and i < len(lst)) else None

    an = info.get("asset_notes") or []
    _info_block(ws, 2, 4, 5, info.get("asset_title") or "기초자산", [
        {"label": "대출실행일", "value": am["start"], "nf": DATEF, "note": _note(an, 0)},
        {"label": "차주", "value": info.get("borrower") or "", "note": _note(an, 1)},
        {"label": "대출금액 (원)", "value": am["amount"], "nf": MONEY, "note": _note(an, 2)},
        {"label": "대출금리", "value": am["rate"], "nf": RATE_INFO, "note": _note(an, 3)},
        {"label": "참여수수료", "value": info.get("part_rate"), "nf": PCT2, "note": _note(an, 4)},
        {"label": "이자지급일",
         "value": info.get("asset_pay_text") or pay_text(am["pay_type"], am["rules"]),
         "note": _note(an, 5)},
        {"label": "만기일", "value": am["mat"], "nf": DATEF, "note": _note(an, 6)},
    ])

    binfo = info.get("bonds") or []
    for k in range(nb):
        bi = binfo[k] if k < len(binfo) else {}
        b = bonds[k]
        base = BOND_BASE[k]
        fee_mode = bi.get("fee_mode", "rate")
        fee_val = bi.get("fee_amount") if fee_mode == "amount" else bi.get("fee_rate")
        bn = bi.get("notes") or []
        _info_block(ws, base, base + 2, base + 3,
                    bi.get("title") or bond_label(k + 1, nb), [
            {"label": "사모사채 발행일", "value": b["start"], "nf": DATEF, "note": _note(bn, 0)},
            {"label": "발행 유형", "value": bi.get("issue_type") or "", "note": _note(bn, 1)},
            {"label": "사모사채 발행금액(원)", "value": b["amount"], "nf": MONEY, "note": _note(bn, 2)},
            {"label": "사모사채 발행금리", "value": b["rate"], "nf": RATE_INFO, "note": _note(bn, 3)},
            {"label": "사모사채 인수수수료(원)", "value": fee_val,
             "nf": MONEY if fee_mode == "amount" else PCT2, "note": _note(bn, 4)},
            {"label": "이자지급일", "value": bi.get("pay_text") or "", "note": _note(bn, 5)},
            {"label": "만기일", "value": bi.get("mat"), "nf": DATEF, "note": _note(bn, 6)},
        ])

    # ── 지급날짜 축 ──
    keys = sorted({p.pay for p in asset}
                  | {b["start"] for b in bonds}
                  | {p.pay for b in bonds for p in b["periods"]})
    row_of = {d: START_ROW + i for i, d in enumerate(keys)}
    last_row = START_ROW + len(keys) - 1
    total_row = last_row + 1

    # ── 헤더 ──
    ws.merge_cells("B12:B13")
    _set(ws, "B12", "지급날짜", bold=True); _box(ws, "B12")
    def _group(base, text):
        ws.merge_cells(f"{_L(base)}12:{_L(base + SEG_W - 1)}12")
        _set(ws, f"{_L(base)}12", text, bold=True, h="none")
        # 병합된 칸 전체에 테두리를 줘야 바깥 선이 끊기지 않는다
        for i in range(SEG_W):
            _box(ws, f"{_L(base + i)}12")
    _group(ASSET_BASE, "▶ 기초자산 이자지급 스케줄 (Cash-in)")
    for k in range(nb):
        _group(BOND_BASE[k], "▶ %s 이자지급 스케줄 (Cash-out)" % bond_label(k + 1, nb))

    WRAP = {2, 4, 5, 6}
    def head(base, labels):
        for i, t in enumerate(labels):
            c = _L(base + i) + "13"
            _set(ws, c, t, bold=True, wrap=(i in WRAP)); _box(ws, c)
    head(ASSET_BASE, ASSET_HEAD)
    for k in range(nb):
        head(BOND_BASE[k], BOND_HEAD)
    if use_fee:
        F = _L(addfee_col(nb))
        _set(ws, F + "12", "추가자산관리수수료", bold=True, wrap=True); _box(ws, F + "12")
        _set(ws, F + "13", "금액(vat포함)", bold=True, wrap=True); _box(ws, F + "13")

    # (단위 : 원)
    _set(ws, _L(last_bond_col(nb)) + "11", "(단위 : 원)", h="right")

    # ── 지급날짜 칸 ──
    for d, r in row_of.items():
        _set(ws, f"B{r}", d, bold=True, nf=DATEF); _box(ws, f"B{r}")

    # ── 기초자산 ──
    AC = seg_cols(ASSET_BASE)
    C, D, E, F, G, H, I = (_L(AC["start"]), _L(AC["end"]), _L(AC["pay"]),
                           _L(AC["days"]), _L(AC["rate"]), _L(AC["fee"]), _L(AC["int"]))
    for idx, p in enumerate(asset):
        r = row_of[p.pay]
        _set(ws, f"{C}{r}", p.start, nf=DATEF); _box(ws, f"{C}{r}")
        _set(ws, f"{D}{r}", p.end, nf=DATEF); _box(ws, f"{D}{r}")
        _set(ws, f"{E}{r}", p.pay, nf=DATEF); _box(ws, f"{E}{r}")
        _set(ws, f"{F}{r}", f"={D}{r}-{C}{r}", nf=NUM); _box(ws, f"{F}{r}")
        _set(ws, f"{G}{r}", "=$D$6", nf=PCT2); _box(ws, f"{G}{r}")
        if idx == 0 and info.get("part_rate") is not None:
            _set(ws, f"{H}{r}", "=$D$5*$D$7", nf=NUM)
        else:
            _set(ws, f"{H}{r}", nf=MONEY)
        _box(ws, f"{H}{r}")
        _set(ws, f"{I}{r}", f"=ROUNDDOWN($D$5*{G}{r}*{F}{r}/365,0)", nf=MONEY); _box(ws, f"{I}{r}")

    # ── 사모사채 ──
    for k in range(nb):
        b = bonds[k]
        cb = seg_cols(BOND_BASE[k])
        bC, bD, bE, bF, bG, bH, bI = (_L(cb["start"]), _L(cb["end"]), _L(cb["pay"]),
                                      _L(cb["days"]), _L(cb["rate"]), _L(cb["fee"]), _L(cb["int"]))
        P = _L(BOND_BASE[k] + 2)          # 정보블록 값 열 (L / S / Z)
        bi = (info.get("bonds") or [{}] * nb)[k] if k < len(info.get("bonds") or []) else {}

        base_r = row_of[b["start"]]
        _set(ws, f"{bC}{base_r}", f"={P}3", nf=DATEF); _box(ws, f"{bC}{base_r}")
        fee_mode = bi.get("fee_mode", "rate")
        given = (bi.get("fee_amount") is not None) if fee_mode == "amount" else (bi.get("fee_rate") is not None)
        if given:
            _set(ws, f"{bH}{base_r}",
                 f"=${P}$7" if fee_mode == "amount" else f"=${P}$7*${P}$5", nf=MONEY)
        else:
            _set(ws, f"{bH}{base_r}", nf=MONEY)
        _box(ws, f"{bH}{base_r}")
        _set(ws, f"{bI}{base_r}", 0, nf=MONEY); _box(ws, f"{bI}{base_r}")

        for p in b["periods"]:
            r = row_of[p.pay]
            _set(ws, f"{bC}{r}", p.start, nf=DATEF); _box(ws, f"{bC}{r}")
            _set(ws, f"{bD}{r}", p.end, nf=DATEF); _box(ws, f"{bD}{r}")
            _set(ws, f"{bE}{r}", p.pay, nf=DATEF); _box(ws, f"{bE}{r}")
            _set(ws, f"{bF}{r}", f"={bD}{r}-{bC}{r}", nf=NUM); _box(ws, f"{bF}{r}")
            _set(ws, f"{bG}{r}", f"=${P}$6", nf=PCT2); _box(ws, f"{bG}{r}")
            _set(ws, f"{bH}{r}", nf=MONEY); _box(ws, f"{bH}{r}")
            _set(ws, f"{bI}{r}", f"=ROUNDDOWN(${P}$5*{bG}{r}*{bF}{r}/365,0)", nf=MONEY)
            _box(ws, f"{bI}{r}")

    # ── 추가자산관리수수료 ──
    #   자금판(후)와 같이 '값'이 아니라 '수식'으로 넣는다.
    #   = 기초자산 이자셀 − 사모사채 이자셀들   (구간 수가 같을 때만)
    if use_fee:
        FL = _L(addfee_col(nb))
        manual = sched.get("addfee_manual") or {}
        valid = [b for b in bonds if b["periods"]]
        same_len = bool(asset) and bool(valid) and all(
            len(b["periods"]) == len(asset) for b in valid)
        if same_len:
            int_cols = [_L(seg_cols(BOND_BASE[k])["int"]) for k in range(nb)]
            for i, ap in enumerate(asset):
                r = row_of[valid[0]["periods"][i].pay]
                if ap.pay in manual or valid[0]["periods"][i].pay in manual:
                    v = manual.get(valid[0]["periods"][i].pay, manual.get(ap.pay))
                    _set(ws, f"{FL}{r}", v, nf=MONEY)
                else:
                    minus = "-".join(
                        "%s%d" % (int_cols[k], row_of[bonds[k]["periods"][i].pay])
                        for k in range(nb) if bonds[k]["periods"])
                    _set(ws, f"{FL}{r}", f"={I}{row_of[ap.pay]}-{minus}", nf=MONEY)
                _box(ws, f"{FL}{r}")

    # ── 빈 칸도 테두리 ──
    lastc = addfee_col(nb) if use_fee else last_bond_col(nb)
    for r in range(START_ROW, total_row):
        for c in range(2, lastc + 1):
            coord = f"{_L(c)}{r}"
            cell = ws.cell(row=r, column=c)
            if cell.border is None or cell.border.top is None or cell.border.top.style is None:
                _box(ws, coord)
            if cell.font is None or cell.font.name != FONT:
                _set(ws, coord)

    # ── 합계행 ──
    ws.merge_cells(f"B{total_row}:{E}{total_row}")
    _set(ws, f"B{total_row}", "합 계", bold=True); _box(ws, f"B{total_row}")

    def sum_col(ci, nf):
        L = _L(ci)
        _set(ws, f"{L}{total_row}", f"=SUM({L}{START_ROW}:{L}{last_row})",
             bold=True, nf=nf, h="none" if nf == MONEY else "center")
        _box(ws, f"{L}{total_row}")

    sum_col(AC["days"], NUM)
    _set(ws, f"{G}{total_row}", nf=PCT2, bold=True); _box(ws, f"{G}{total_row}")
    sum_col(AC["fee"], MONEY)
    sum_col(AC["int"], MONEY)
    for k in range(nb):
        cb = seg_cols(BOND_BASE[k])
        sum_col(cb["days"], NUM)
        sum_col(cb["fee"], MONEY)
        sum_col(cb["int"], MONEY)
    if use_fee:
        sum_col(addfee_col(nb), MONEY)
    for c in range(2, lastc + 1):
        coord = f"{_L(c)}{total_row}"
        cell = ws.cell(row=total_row, column=c)
        if cell.border is None or cell.border.top is None or cell.border.top.style is None:
            _box(ws, coord)
        # (자금판(후) 는 합계행 빈 칸에 글꼴을 주지 않는다 — 그대로 맞춘다)

    # ── 색 ──
    _fill_rng(ws, 2, 12, 2, 13, F_GRAY, C_NAVY)
    _fill_rng(ws, 2, 2, 6, 2, F_ASSET, "FFFFFFFF")
    _fill_rng(ws, ASSET_BASE, 12, ASSET_BASE + SEG_W - 1, 13, F_ASSET, "FFFFFFFF")
    for k in range(nb):
        base = BOND_BASE[k]
        _fill_rng(ws, base, 2, base + 4, 2, F_BONDS[k], "FFFFFFFF")
        _fill_rng(ws, base, 12, base + SEG_W - 1, 13, F_BONDS[k], "FFFFFFFF")
    if use_fee:
        c = addfee_col(nb)
        _fill_rng(ws, c, 12, c, 13, F_FEE, C_NAVY)
    _fill_rng(ws, 2, 3, 3, 9, F_GRAY)
    for k in range(nb):
        _fill_rng(ws, BOND_BASE[k], 3, BOND_BASE[k] + 1, 9, F_GRAY)
    _fill_rng(ws, 2, total_row, lastc, total_row, F_GRAY)

    # ── 세그먼트 구분선 (12행 ~ 합계행) ──
    #   ※ openpyxl 은 저장할 때 병합범위의 테두리를 '왼쪽 위 칸' 것으로 다시 칠한다.
    #     그래서 병합범위 안의 칸에 선을 주면 저장하면서 사라진다.
    #     병합된 자리면 그 범위의 왼쪽 위 칸에 준다(= 범위 바깥선이 그 칸을 따라간다).
    def _anchor_cell(r, c):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                return rng.min_row, rng.min_col
        return r, c

    def vline(ci, side):
        if ci < 2:
            return
        for r in range(12, total_row + 1):
            rr, cc = _anchor_cell(r, ci)
            cell = ws.cell(row=rr, column=cc)
            b = cell.border
            kw = {"left": b.left, "right": b.right, "top": b.top, "bottom": b.bottom}
            kw[side] = _MED
            cell.border = Border(**kw)

    def hline(c1, c2, r, side):
        for c in range(c1, c2 + 1):
            rr, cc = _anchor_cell(r, c)
            cell = ws.cell(row=rr, column=cc)
            b = cell.border
            kw = {"left": b.left, "right": b.right, "top": b.top, "bottom": b.bottom}
            kw[side] = _MED
            cell.border = Border(**kw)

    starts = [2, ASSET_BASE] + BOND_BASE[:nb]
    if use_fee:
        starts.append(addfee_col(nb))
    for c in starts:
        vline(c, "left"); vline(c - 1, "right")
    vline(lastc, "right")

    # 표 전체의 위·아래도 굵게 — 기초자산 / 1·2·3회가 각각 상자로 보이게
    hline(2, lastc, 12, "top")
    hline(2, lastc, total_row, "bottom")

    return {"total_row": total_row, "row_of": row_of, "last_row": last_row,
            "asset_int_col": I, "bond_int_cols": [_L(seg_cols(BOND_BASE[k])["int"]) for k in range(nb)],
            "addfee_col": _L(addfee_col(nb)) if use_fee else None,
            "asset_pay_col": E}


# ─────────────────────────────────────────────
# 후순위대여 표 — 이자 스케줄 합계행 아래 3칸 띄우고
# ─────────────────────────────────────────────
def write_wht(ws, meta: dict, asset: list, rate_pct: float, local_pct: float):
    top = meta["total_row"] + 4
    _set(ws, f"B{top}", "후순위대여", bold=True, h="left")
    hr = top + 1
    for i, t in enumerate(["이자지급일", "이자금액(세전)", "원천세", "지방세", "합계"]):
        c = _L(2 + i) + str(hr)
        _set(ws, c, t, bold=True, fill=F_GRAY); _box(ws, c)

    first = hr + 1
    payL, intL = meta["asset_pay_col"], meta["asset_int_col"]
    for i, p in enumerate(asset):
        r = first + i
        src = meta["row_of"][p.pay]
        _set(ws, f"B{r}", f"={payL}{src}", nf=DATEF); _box(ws, f"B{r}")
        _set(ws, f"C{r}", f"={intL}{src}", nf=MONEY, h="right"); _box(ws, f"C{r}")
        _set(ws, f"D{r}", f"=ROUNDDOWN(C{r}*{rate_pct}%,-1)", nf=MONEY, h="right"); _box(ws, f"D{r}")
        _set(ws, f"E{r}", f"=ROUNDDOWN(D{r}*{local_pct}%,-1)", nf=MONEY, h="right"); _box(ws, f"E{r}")
        _set(ws, f"F{r}", f"=E{r}+D{r}", nf=MONEY, h="right"); _box(ws, f"F{r}")

    last = first + len(asset) - 1
    sr = last + 1
    _set(ws, f"B{sr}", "합 계", bold=True, fill=F_GRAY); _box(ws, f"B{sr}")
    for L in ("C", "D", "E", "F"):
        _set(ws, f"{L}{sr}", f"=SUM({L}{first}:{L}{last})", bold=True, nf=MONEY,
             h="right", fill=F_GRAY)
        _box(ws, f"{L}{sr}")
    return {"first": first, "sum_row": sr}
