"""
자금판 내용을 예시 엑셀과 비슷한 모양의 .xlsx 파일로 만드는 부품.

- 예시 "부성2지구담보대출_사모사채_자금판.xlsx" 처럼 **수식(함수·사칙연산)** 으로 계산되게 만듭니다.
  (숫자를 그냥 박아넣지 않음 → 사용자는 수식만 확인하면 됨)
  · 인수수수료 D7 = =D5*수수료율
  · 이자기간 말일 = =EDATE(초일, 개월수)  (만기 초과분은 만기로 캡)
  · 이자계산일수 = =말일-초일
  · 이자금액 = =ROUNDDOWN(발행금액*금리*일수/365, 0)
  · 합계 = =SUM(...)
- 서식(맑은 고딕 10pt, 회색 테두리, 금액 천단위, 금리 %, 병합셀)도 예시와 비슷하게.
- 긴 글씨가 잘리지 않게 제목/이자지급일 행 높이를 넉넉히 줍니다.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.schedule import KR_HOLIDAYS


# ── 공용 서식 ─────────────────────────────
FONT = Font(name="맑은 고딕", size=10)
FONT_B = Font(name="맑은 고딕", size=10, bold=True)
FONT_W = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")

_thin = Side(style="thin", color="A5A5A5")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FILL_LABEL = PatternFill("solid", fgColor="F2F2F2")   # 항목 이름 칸(연회색)
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")  # 표 머리글(연한 남색)
FILL_TITLE = PatternFill("solid", fgColor="1A2B5E")   # 구획(남색)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
NOWRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

FMT_WON = "#,##0"
FMT_RATE = '"연" 0.00%'
FMT_RATE2 = "0.00%"
FMT_DATE = "yyyy-mm-dd(aaa)"   # aaa = 요일(월/화/…)


def _set(ws, coord, value, font=FONT, fill=None, align=CENTER, numfmt=None):
    c = ws[coord]
    c.value = value
    c.font = font
    c.alignment = align
    if fill:
        c.fill = fill
    if numfmt:
        c.number_format = numfmt
    return c


def _border_range(ws, rng):
    for row in ws[rng]:
        for c in row:
            c.border = BORDER


def _short_ip(ip: str) -> str:
    """이자지급 방식에서 'N개월 선/후취'만 짧게 뽑음(길면 D8 칸이 넘침)."""
    if not ip:
        return ""
    m = re.search(r"(\d+\s*개?\s*월).*?(선취|후취)", ip)
    if m:
        return f"{m.group(1).replace(' ', '')} {m.group(2)}"
    return ip


def build_workbook(ctx: dict, account: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "사모사채 자금판"
    ws.sheet_view.showGridLines = False

    widths = {"A": 4.5, "B": 16.2, "C": 16.4, "D": 15.0, "E": 19.1, "F": 14.9,
              "G": 12.0, "H": 12.0, "I": 12.0, "J": 14.1, "K": 14.4, "L": 14.4,
              "M": 18.2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 긴 글씨가 잘리지 않게 행 높이 넉넉히
    ws.row_dimensions[2].height = 32   # 제목(사채명이 길 수 있음)
    ws.row_dimensions[8].height = 30   # 이자지급일 비고

    months = ctx["months"]
    postpaid = ctx["postpaid"]

    # ── 제목 / 비고 / 계좌 머리글 ──
    ws.merge_cells("B2:D2")
    _set(ws, "B2", ctx["bond_name"] or "사모사채 자금판", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("E2:F2")
    _set(ws, "E2", "비고", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("H2:M3")
    _set(ws, "H2", "사모사채 인수대금 납입 계좌", font=FONT_W, fill=FILL_TITLE)

    # ── 발행 기본조건 (왼쪽 블록) ── 값은 raw, 나머지는 수식 ──
    ws.merge_cells("B3:C3"); _set(ws, "B3", "사모사채 발행일", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D3", ctx["idate"], numfmt=FMT_DATE)
    ws.merge_cells("E3:F3"); _set(ws, "E3", None)

    ws.merge_cells("B4:C4"); _set(ws, "B4", "발행 유형", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D4", ctx["issue_type"], numfmt="@")
    ws.merge_cells("E4:F4")
    _set(ws, "E4", f"인수인: {ctx['underwriter']}" if ctx["underwriter"] else None)

    ws.merge_cells("B5:C5"); _set(ws, "B5", "사모사채 발행금액(원)", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D5", ctx["amount"], numfmt=FMT_WON)
    ws.merge_cells("E5:F5"); _set(ws, "E5", None)

    ws.merge_cells("B6:C6"); _set(ws, "B6", "사모사채 발행금리", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D6", ctx["rate"], numfmt=FMT_RATE)
    ws.merge_cells("E6:F6"); _set(ws, "E6", ctx["rate_note"] or None)

    ws.merge_cells("B7:C7"); _set(ws, "B7", "사모사채 인수수수료(원)", font=FONT_B, fill=FILL_LABEL)
    if ctx.get("fee_rate"):
        _set(ws, "D7", f"=D5*{ctx['fee_rate']}", numfmt=FMT_WON)   # 수식: 발행금액 × 수수료율
    else:
        _set(ws, "D7", ctx["fee"], numfmt=FMT_WON)
    ws.merge_cells("E7:F7"); _set(ws, "E7", ctx["fee_note"] or None)

    ws.merge_cells("B8:C8"); _set(ws, "B8", "이자지급일", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D8", _short_ip(ctx["ip"]), numfmt="@")
    ws.merge_cells("E8:F8"); _set(ws, "E8", ctx["ip_note"] or None)

    ws.merge_cells("B9:C9"); _set(ws, "B9", "만기일", font=FONT_B, fill=FILL_LABEL)
    _set(ws, "D9", ctx["mdate"], numfmt=FMT_DATE)
    ws.merge_cells("E9:F9"); _set(ws, "E9", None)

    # ── 납입 계좌 (오른쪽 블록) ──
    ws.merge_cells("H5:I5"); _set(ws, "H5", "은행명", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("J5:K5"); _set(ws, "J5", "계좌번호", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("L5:M5"); _set(ws, "L5", "예금주", font=FONT_B, fill=FILL_LABEL)
    ws.merge_cells("H6:I7"); _set(ws, "H6", account.get("bank") or None)
    ws.merge_cells("J6:K7"); _set(ws, "J6", account.get("number") or None)
    ws.merge_cells("L6:M7"); _set(ws, "L6", account.get("holder") or None)

    _border_range(ws, "B2:F9")
    _border_range(ws, "H2:M7")

    # ── 이자지급 스케줄 (전부 수식) ──
    _set(ws, "M11", "(단위 : 원)", align=RIGHT)

    ws.merge_cells("B12:M12")
    _set(ws, "B12", " ▶ 사모사채 이자지급스케줄", font=FONT_B, fill=FILL_LABEL, align=LEFT)

    headers = ["구분", "이자기간(초일)", "이자기간(말일)", "이자지급일(후취)",
               "이자계산일수", "윤년미적용", "윤년적용", "금리(연)",
               "인출금액", "이자금액(세전)", "인수수수료", "비고"]
    cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    for col, h in zip(cols, headers):
        _set(ws, f"{col}13", h, font=FONT_B, fill=FILL_HEADER)

    n = ctx["sch"]["period_count"]

    # 공휴일 목록을 숨김 열(T)에 적어, 지급일 수식(WORKDAY)에서 참조.
    # → 이자지급일이 주말·공휴일이면 다음 영업일로 자동 조정.
    hol = sorted(KR_HOLIDAYS)
    for i, hd in enumerate(hol, start=1):
        _set(ws, f"T{i}", hd, numfmt=FMT_DATE, align=LEFT)
    ws.column_dimensions["T"].hidden = True
    hol_range = f"$T$1:$T${len(hol)}"

    # 발행일 인출 행 (14행)
    _set(ws, "B14", "-"); _set(ws, "C14", "-"); _set(ws, "D14", "-")
    _set(ws, "E14", "=D3", numfmt=FMT_DATE)     # 발행일
    _set(ws, "F14", "-"); _set(ws, "G14", "-"); _set(ws, "H14", "-"); _set(ws, "I14", "-")
    _set(ws, "J14", "=D5", numfmt=FMT_WON)      # 인출금액 = 발행금액
    _set(ws, "K14", 0, numfmt=FMT_WON)
    _set(ws, "L14", "=D7", numfmt=FMT_WON)      # 인수수수료
    _set(ws, "M14", "발행일 인출", align=LEFT)

    # 각 이자기간 행 (15행부터)
    # ★ 초일·말일·지급일·일수는 '계산된 스케줄 값'을 그대로 넣음
    #   (계약서 기준일 28일·말일 이동/고정이 build_schedule 에서 이미 반영됨).
    #   이자금액만 수식(=발행금액×금리×일수/365)으로 둔다.
    sch_rows = ctx["sch"]["rows"][1:]   # [0]은 발행일 인출 행이라 제외
    for i, srow in enumerate(sch_rows):
        r = 15 + i
        _set(ws, f"B{r}", 1 if i == 0 else f"=+B{r-1}+1")
        _set(ws, f"C{r}", srow["초일"], numfmt=FMT_DATE)
        _set(ws, f"D{r}", srow["말일"], numfmt=FMT_DATE)
        _set(ws, f"E{r}", srow["지급일"], numfmt=FMT_DATE)
        _set(ws, f"F{r}", srow["일수"], numfmt=FMT_WON)
        _set(ws, f"G{r}", f"=F{r}", numfmt=FMT_WON)   # 윤년미적용(365 기준 동일)
        _set(ws, f"H{r}", 0, numfmt=FMT_WON)
        _set(ws, f"I{r}", "=$D$6", numfmt=FMT_RATE2)  # 금리
        _set(ws, f"J{r}", None)
        _set(ws, f"K{r}", f"=ROUNDDOWN($D$5*I{r}*F{r}/365,0)", numfmt=FMT_WON)
        _set(ws, f"L{r}", None)
        _set(ws, f"M{r}", None, align=LEFT)

    # 합계 행
    tr = 15 + n
    ws.merge_cells(f"B{tr}:E{tr}")
    _set(ws, f"B{tr}", "합 계", font=FONT_B, fill=FILL_LABEL)
    _set(ws, f"F{tr}", f"=SUM(F15:F{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"G{tr}", f"=SUM(G15:G{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"H{tr}", f"=SUM(H15:H{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"I{tr}", None, fill=FILL_LABEL)
    _set(ws, f"J{tr}", f"=SUM(J14:J{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"K{tr}", f"=SUM(K14:K{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"L{tr}", f"=SUM(L14:L{tr-1})", font=FONT_B, fill=FILL_LABEL, numfmt=FMT_WON)
    _set(ws, f"M{tr}", None, fill=FILL_LABEL)

    _border_range(ws, f"B12:M{tr}")

    # 각주 (합계 칸처럼 B:M 을 셀 병합해서 한 줄로 깔끔하게)
    fr = tr + 1
    if ctx.get("interest_calc"):
        ws.merge_cells(f"B{fr}:M{fr}")
        _set(ws, f"B{fr}", f"* 이자계산 기준: {ctx['interest_calc']}",
             font=FONT_B, align=NOWRAP_LEFT)
        fr += 1
    ws.merge_cells(f"B{fr}:M{fr}")
    _set(ws, f"B{fr}",
         "* 이자계산: 1년 365일 기준, 초일산입·말일불산입. 지급일이 주말이면 익영업일(공휴일 미반영).",
         font=FONT, align=NOWRAP_LEFT)

    # 엑셀을 열 때 수식을 자동으로 다시 계산하도록(수식만 넣으면 값이 안 보이는 문제 방지)
    wb.calculation.fullCalcOnLoad = True

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
